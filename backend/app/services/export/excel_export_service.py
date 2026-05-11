from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


class ExcelExportService:
    @staticmethod
    def write_email_drafts(rows: list[dict[str, object]], file_path: Path) -> None:
        file_path.parent.mkdir(parents=True, exist_ok=True)
        dataframe = pd.DataFrame(rows)

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            dataframe.to_excel(writer, sheet_name="Email Drafts", index=False)
            worksheet = writer.sheets["Email Drafts"]
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions

            header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font

            for column_cells in worksheet.columns:
                max_length = 0
                column_index = column_cells[0].column
                column_letter = get_column_letter(column_index)
                for cell in column_cells:
                    value = "" if cell.value is None else str(cell.value)
                    max_length = max(max_length, len(value))

                worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)

            worksheet.column_dimensions["I"].width = 36
            worksheet.column_dimensions["J"].width = 60
